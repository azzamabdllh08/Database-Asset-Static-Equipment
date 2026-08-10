from pathlib import Path
import json
import re
import openpyxl

INPUT = Path("Input RBI.xlsx")
OUTPUT = Path("data.js")

# Database is read-only with respect to RBI logic: values are copied from Excel.
ALIASES = {
    "tag": ["Line No.", "Line No", "Tag No.", "Tag No", "Equipment No.", "Equipment No"],
    "name": ["Item Description", "Equipment Name", "Equipment", "Description"],
    "type": ["Equipment Type", "Type"],
    "area": ["P/F", "Area", "Location Area"],
    "location": ["Location", "Originating"],
    "service": ["Service Fluid", "Service", "Process Service"],
    "system": ["System Name", "System"],
    "material": ["Material", "Material of Construction", "MOC"],
    "designPressure": ["DP (psig)", "DP\\n(psig)", "Design Pressure"],
    "designTemperature": ["DT (0F)", "DT \\n(0F)", "Design Temperature"],
    "operatingPressure": ["Max. OP (psig)", "Max. OP\\n(psig)", "Operating Pressure"],
    "operatingTemperature": ["Max. OT (0F)", "Max. OT \\n(0F)", "Operating Temperature"],
    "nps": ["NPS"],
    "schedule": ["SCH."],
    "nominalThickness": ["Nom. THK.\\n(mm)", "Nom. THK. (mm)", "Nominal Thickness"],
    "corrosionAllowance": ["CA \\n(mm)", "CA (mm)", "Corrosion Allowance", "CA"],
    "assessmentDate": ["Assessment Date"],
    "inspectionDate": ["Last Insp. Date", "Inspection Date"],
    "inspectionMethod": ["Inspection Method"],
    "visualFinding": ["Visual Finding"],
    "thicknessMeasurement": ["Thickness Measurement"],
    "internalIssue": ["Internal issue\\n(Yes/No)", "Internal issue (Yes/No)"],
    "externalIssue": ["External issue\\n(Yes/No)", "External issue (Yes/No)"],
    "repair": ["Repair\\n(Yes/No)", "Repair (Yes/No)"],
    "repairDescription": ["Repair Description"],
    "lf1": ["LF1"], "lf3": ["LF3"], "lf4": ["LF4"], "lf5": ["LF5"],
    "lf6": ["LF6"], "lf7": ["LF7"],
    "cf5": ["CF5"], "cf6": ["CF6"], "cf7": ["CF7"], "cf8": ["CF8"], "cf9": ["CF9"],
    "damageMechanismInternal": ["Potential DM's\\nInternal", "Potential DM's Internal"],
    "damageMechanismExternal": ["Potential DM's\\nExternal", "Potential DM's External"],
    "elInternal": ["EL \\nInternal", "EL Internal"],
    "elExternal": ["EL \\nExternal", "EL External"],
    "minEL": ["Min EL (Months)"],
    "risk1AP": ["Risk \\n(1AP)", "Risk (1AP)", "Risk 1AP"],
    "risk2AP": ["Risk \\n(2AP)", "Risk (2AP)", "Risk 2AP"],
    "risk3AP": ["Risk \\n(3AP)", "Risk (3AP)", "Risk 3AP"],
    "rli": ["RLI \\n(Months)", "RLI (Months)", "RLI"],
    "criticality1AP": ["Criticality 1AP", "Criticality"],
    "rliDueDate": ["RLI Due Date\\nmm/dd/yyyy", "RLI Due Date"],
    "inspectionDueDate": ["Inspection Due Date\\nmm/dd/yyyy", "Inspection Due Date"],
    "inspectionScope": ["Inspection Scope\\nyyyy", "Inspection Scope"],
    "note": ["Note"], "remarks": ["Remarks", "Remark"],
}


def norm(v):
    return re.sub(r"\\s+", " ", str(v or "").strip().lower())


def clean(v):
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return v


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        normalized = {norm(v) for v in vals if v is not None}
        if "line no." in normalized or "tag no." in normalized or "equipment name" in normalized:
            return r
    return 1


def find_column(headers, aliases):
    normalized = {norm(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        if norm(alias) in normalized:
            return normalized[norm(alias)]
    return None

if not INPUT.exists():
    raise SystemExit(f"Missing {INPUT}")

wb = openpyxl.load_workbook(INPUT, data_only=True, read_only=True)
ws = wb["AssReg GUNDIH-PP"] if "AssReg GUNDIH-PP" in wb.sheetnames else wb[wb.sheetnames[0]]
header_row = find_header_row(ws)
rows = list(ws.iter_rows(min_row=header_row, values_only=True))
headers = list(rows[0])
columns = {k: find_column(headers, aliases) for k, aliases in ALIASES.items()}

assets = []
for row in rows[1:]:
    if not any(v not in (None, "") for v in row):
        continue
    def val(key):
        idx = columns.get(key)
        return clean(row[idx]) if idx is not None and idx < len(row) else ""

    tag = str(val("tag")).strip()
    if not tag:
        continue

    item = {k: val(k) for k in ALIASES}
    item["risk"] = item["risk1AP"]  # dashboard default view = 1AP; all three AP values remain stored
    item["rbiStatus"] = item["criticality1AP"]
    assets.append(item)

OUTPUT.write_text(
    "const ASSETS = " + json.dumps(assets, ensure_ascii=False, indent=2, default=str) + ";\n\n" +
    "const INSPECTIONS = [];\n",
    encoding="utf-8"
)

print(f"Generated {OUTPUT} from {INPUT}: {len(assets)} records")
print("RBI/Risk/Corrosion calculations are NOT performed; values are copied from Excel.")
