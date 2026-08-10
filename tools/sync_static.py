from __future__ import annotations

import argparse
import json
import re
import subprocess
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ALIASES = {
    "category": ["Equipment Category", "EquipmentCategory", "Category"],
    "wilayahKerja": ["Wilayah Kerja", "Wilayah", "Working Area", "Work Area"],
    "tag": ["Tag No", "Tag No.", "Tag Number", "Equipment Tag", "Tag", "Equipment No"],
    "name": ["Deskripsi Peralatan", "Equipment Description", "Description", "Equipment Name", "Equipment", "Name"],
    "objectType": ["Object Type", "ObjectType"],
    "type": ["Object Type", "ObjectType", "Equipment Type", "Type", "Equipment Type Description"],
    "area": ["Area", "Location", "Plant Area", "Unit", "Zone"],
    "service": ["Service", "Process Service", "Fluid Service", "Service Description"],
    "material": ["Material", "Material Specification", "Material Spec", "Material Grade"],
    "integrityStatus": ["Integrity Status", "Integrity status AZ11APS", "Integrity Status AZ11APS"],
    "risk": ["Risk 1AP", "Risk 1 AP", "1AP", "1 AP", "Risk", "Risk Ranking", "Risk Rank", "RBI Risk"],
    "risk1AP": ["Risk 1AP", "Risk 1 AP", "1AP", "1 AP"],
    "risk2AP": ["Risk 2AP", "Risk 2 AP", "2AP", "2 AP"],
    "risk3AP": ["Risk 3AP", "Risk 3 AP", "3AP", "3 AP"],
    "damageMechanism": ["Damage Mechanism", "DamageMechanism", "DM"],
    "corrosionRate": ["Corrosion Rate", "CorrosionRate", "CR"],
    "currentThickness": ["Current Thickness", "CurrentThickness", "Measured Thickness", "UT Thickness"],
    "rbiStatus": ["RBI Status", "RBIStatus", "Assessment Status"],
    "date": ["Inspection Date", "Last Inspection Date", "InspectionDate", "Date"],
    "method": ["Inspection Method", "Method", "Inspection Type"],
    "finding": ["Finding", "Findings", "Inspection Finding"],
    "remarks": ["Remarks", "Remark", "Inspection Remarks", "Notes"],
}


def clean_header(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def value_for(row, header_map, aliases):
    for alias in aliases:
        key = clean_header(alias)
        if key in header_map:
            return row[header_map[key]]
    return None


def serialise(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def build_records(rows, headers):
    header_map = {clean_header(h): i for i, h in enumerate(headers) if h is not None}
    if not any(clean_header(a) in header_map for a in ALIASES["category"]):
        raise RuntimeError(
            "Kolom Equipment Category tidak ditemukan. Header yang tersedia: "
            + ", ".join(str(h) for h in headers if h is not None)
        )

    assets = []
    inspections = []
    seen_tags = set()

    for row in rows:
        raw_category = value_for(row, header_map, ALIASES["category"])
        if str(raw_category or "").strip().lower() != "static":
            continue

        tag = serialise(value_for(row, header_map, ALIASES["tag"]))
        tag_key = str(tag or "").strip().upper()
        if not tag_key or tag_key in seen_tags:
            continue
        seen_tags.add(tag_key)

        asset = {
            "wilayahKerja": serialise(value_for(row, header_map, ALIASES["wilayahKerja"])) or "Unknown",
            "tag": tag,
            "name": serialise(value_for(row, header_map, ALIASES["name"])),
            "objectType": serialise(value_for(row, header_map, ALIASES["objectType"])),
            "type": serialise(value_for(row, header_map, ALIASES["type"])),
            "area": serialise(value_for(row, header_map, ALIASES["area"])),
            "service": serialise(value_for(row, header_map, ALIASES["service"])),
            "material": serialise(value_for(row, header_map, ALIASES["material"])),
            "integrityStatus": serialise(value_for(row, header_map, ALIASES["integrityStatus"])),
            "risk1AP": serialise(value_for(row, header_map, ALIASES["risk1AP"])),
            "risk2AP": serialise(value_for(row, header_map, ALIASES["risk2AP"])),
            "risk3AP": serialise(value_for(row, header_map, ALIASES["risk3AP"])),
            "risk": serialise(value_for(row, header_map, ALIASES["risk1AP"])),
            "damageMechanism": serialise(value_for(row, header_map, ALIASES["damageMechanism"])),
            "corrosionRate": serialise(value_for(row, header_map, ALIASES["corrosionRate"])),
            "currentThickness": serialise(value_for(row, header_map, ALIASES["currentThickness"])),
            "rbiStatus": serialise(value_for(row, header_map, ALIASES["rbiStatus"])),
            "equipmentCategory": "Static",
        }
        assets.append(asset)

        inspection_date = value_for(row, header_map, ALIASES["date"])
        method = value_for(row, header_map, ALIASES["method"])
        finding = value_for(row, header_map, ALIASES["finding"])
        remarks = value_for(row, header_map, ALIASES["remarks"])
        if any(v not in (None, "") for v in (inspection_date, method, finding, remarks)):
            inspections.append({
                "tag": tag,
                "date": serialise(inspection_date),
                "method": serialise(method),
                "finding": serialise(finding),
                "remarks": serialise(remarks),
            })

    return assets, inspections


def find_header_in_sheet(ws):
    category_aliases = {clean_header(a) for a in ALIASES["category"]}
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), start=1):
        row_values = list(row)
        header_keys = {clean_header(v) for v in row_values if v is not None}
        if category_aliases.intersection(header_keys):
            return row_number, row_values
    return None, None


def read_excel(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)

    preferred = next((ws for ws in wb.worksheets if clean_header(ws.title) == "detail"), None)
    if preferred is not None:
        row_number, headers = find_header_in_sheet(preferred)
        if row_number is not None:
            remaining = preferred.iter_rows(min_row=row_number + 1, values_only=True)
            return list(remaining), headers, preferred.title

    category_aliases = {clean_header(a) for a in ALIASES["category"]}
    for ws in wb.worksheets:
        row_number, headers = find_header_in_sheet(ws)
        if row_number is not None:
            header_keys = {clean_header(v) for v in headers if v is not None}
            if category_aliases.intersection(header_keys):
                remaining = ws.iter_rows(min_row=row_number + 1, values_only=True)
                return list(remaining), headers, ws.title

    raise RuntimeError(
        "Tidak ditemukan worksheet/baris header yang memiliki kolom Equipment Category "
        "dalam 100 baris pertama."
    )


def slugify(value):
    value = str(value or "Unknown").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value or "unknown"


def count_by(records, key):
    result = {}
    for item in records:
        value = str(item.get(key) or "Unknown")
        result[value] = result.get(value, 0) + 1
    return dict(sorted(result.items(), key=lambda kv: (-kv[1], kv[0].lower())))


def write_database(repo_root: Path, assets, inspections):
    data_root = repo_root / "data"
    regions_root = data_root / "regions"
    regions_root.mkdir(parents=True, exist_ok=True)

    for old in regions_root.glob("*.json"):
        old.unlink()

    region_groups = {}
    for asset in assets:
        region = str(asset.get("wilayahKerja") or "Unknown").strip() or "Unknown"
        region_groups.setdefault(region, []).append(asset)

    region_entries = []
    for region, records in sorted(region_groups.items(), key=lambda kv: kv[0].lower()):
        slug = slugify(region)
        payload = {
            "wilayahKerja": region,
            "count": len(records),
            "locations": count_by(records, "area"),
            "assets": records,
        }
        (regions_root / f"{slug}.json").write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        region_entries.append({
            "name": region,
            "slug": slug,
            "count": len(records),
            "locations": len(count_by(records, "area")),
        })

    recent = assets[:8]
    risk_values = {"4A", "4B", "4C", "4D", "4E", "5A", "5B", "5C", "5D", "5E"}
    manifest = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "totalAssets": len(assets),
        "totalInspections": len(inspections),
        "totalRbi": sum(1 for x in assets if x.get("risk1AP")),
        "highRisk": sum(1 for x in assets if str(x.get("risk1AP") or "") in risk_values),
        "typeCounts": count_by(assets, "objectType"),
        "riskCounts": count_by(assets, "risk1AP"),
        "regions": region_entries,
        "recentAssets": recent,
    }
    data_root.mkdir(parents=True, exist_ok=True)
    (data_root / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (data_root / "inspections.json").write_text(
        json.dumps(inspections, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    output = repo_root / "data.js"
    payload_assets = json.dumps(assets, ensure_ascii=False, separators=(",", ":"))
    payload_inspections = json.dumps(inspections, ensure_ascii=False, separators=(",", ":"))
    output.write_text(
        "// AUTO-GENERATED. Legacy compatibility copy; dashboard does not load this file on startup.\n"
        f"const ASSETS = {payload_assets};\n\nconst INSPECTIONS = {payload_inspections};\n",
        encoding="utf-8",
    )
    return output


def git_push(repo_root: Path, commit_message: str):
    subprocess.run(["git", "-C", str(repo_root), "add", "data.js", "data"], check=True)
    status = subprocess.run(
        ["git", "-C", str(repo_root), "status", "--porcelain"],
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()
    if not status:
        print("Tidak ada perubahan database.")
        return
    subprocess.run(["git", "-C", str(repo_root), "commit", "-m", commit_message], check=True)
    subprocess.run(["git", "-C", str(repo_root), "push", "origin", "main"], check=True)


def main():
    parser = argparse.ArgumentParser(description="Sync Static Equipment dari Asset Register lokal OneDrive")
    parser.add_argument("excel", help="Path lengkap ke Asset Register Excel")
    parser.add_argument("--repo", default=".", help="Folder root repository GitHub")
    parser.add_argument("--push", action="store_true", help="Commit dan push database ke branch main")
    args = parser.parse_args()

    excel = Path(args.excel).expanduser().resolve()
    repo = Path(args.repo).expanduser().resolve()
    if not excel.exists():
        raise SystemExit(f"File Excel tidak ditemukan: {excel}")
    if not (repo / ".git").exists():
        raise SystemExit(f"Folder repository Git tidak ditemukan: {repo}")

    rows, headers, sheet = read_excel(excel)
    assets, inspections = build_records(rows, headers)
    output = write_database(repo, assets, inspections)

    print(f"Worksheet: {sheet}")
    print(f"Static assets: {len(assets):,}")
    print(f"Inspection records: {len(inspections):,}")
    print(f"Wilayah Kerja: {len(set(str(x.get('wilayahKerja') or 'Unknown') for x in assets))}")
    print(f"Database: {output}")

    if args.push:
        git_push(repo, "Sync Static Equipment database by Wilayah Kerja")
        print("Push ke GitHub selesai.")


if __name__ == "__main__":
    main()
